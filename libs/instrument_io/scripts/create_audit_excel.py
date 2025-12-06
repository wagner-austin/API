
import polars as pl
from pathlib import Path
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import PatternFill, Font

def create_audit_excel():
    base_path = Path(r"C:\Users\austi\PROJECTS\UC Irvine\Celia Louise Braun Faiola - FaiolaLab\Notebooks\Emily Truong Notebook")
    inventory_excel_path = base_path / "Chemical_Inventory_List_2025.xlsx"
    standards_excel_path = base_path / "Chemical_Standards_List_2025.xlsx"
    output_path = base_path / "Chemical_Standards_Audit_2025.xlsx"

    # Load Data
    try:
        df_inventory = pl.read_excel(source=inventory_excel_path, engine="openpyxl")
        df_standards = pl.read_excel(source=standards_excel_path, engine="openpyxl")
    except Exception as e:
        print(f"Error loading source files: {e}")
        return

    # Normalize Helper
    def normalize(name):
        if not name or not isinstance(name, str):
            return ""
        return name.strip().lower()

    # Create Dictionaries for fast lookup
    inv_map = {normalize(row["Chemical Name"]): row for row in df_inventory.iter_rows(named=True) if row.get("Chemical Name")}
    std_map = {normalize(row["Chemical Name"]): row for row in df_standards.iter_rows(named=True) if row.get("Chemical Name")}

    # Master set of all unique normalized names
    all_names = set(inv_map.keys()).union(set(std_map.keys()))

    # Build Audit Data
    audit_data = []
    for name_norm in sorted(all_names):
        inv_data = inv_map.get(name_norm, {})
        std_data = std_map.get(name_norm, {})

        # Determine Status
        has_inv = bool(inv_data)
        has_std = bool(std_data)

        if has_inv and has_std:
            status = "OK: Complete"
            name_display = inv_data.get("Chemical Name") # Prefer Inventory Name
        elif has_inv and not has_std:
            status = "Warning: No Standard"
            name_display = inv_data.get("Chemical Name")
        elif not has_inv and has_std:
            status = "Critical: Missing Inventory"
            name_display = std_data.get("Chemical Name")
        else:
            status = "Error"
            name_display = name_norm

        # Merge Sources
        sources = []
        if has_inv:
            sources.append("UCI Chemical Inventory (Risk & Safety)")
        if has_std:
            std_source = std_data.get("Source")
            if std_source:
                sources.append(std_source)
        
        combined_source = "; ".join(sources)

        # Merge Dates
        dates = set()
        if has_inv:
            d = inv_data.get("Date")
            if d: dates.add(str(d))
        if has_std:
            d = std_data.get("Date")
            if d: dates.add(str(d))
        
        combined_date = "; ".join(sorted(dates))

        entry = {
            "Chemical Name": name_display,
            "Status": status,
            "CAS (Inv)": inv_data.get("CAS"),
            "Physical State": inv_data.get("Physical State"),
            "Source": combined_source,
            "Date": combined_date,
            "Details (Std)": std_data.get("Details")
        }
        audit_data.append(entry)

    # Create Excel Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Audit Report"

    # Headers
    headers = ["Chemical Name", "Status", "CAS (Inv)", "Physical State", "Source", "Date", "Details (Std)"]
    for col_idx, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_idx, value=header)

    # Write Data
    last_row = 1
    for row_idx, entry in enumerate(audit_data, 2):
        last_row = row_idx
        ws.cell(row=row_idx, column=1, value=entry["Chemical Name"])
        
        # Status Cell with Conditional Formatting Logic (Manual coloring for now)
        status_cell = ws.cell(row=row_idx, column=2, value=entry["Status"])
        if "OK" in entry["Status"]:
            status_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid") # Green
            status_cell.font = Font(color="006100")
        elif "Warning" in entry["Status"]:
            status_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid") # Yellow
            status_cell.font = Font(color="9C5700")
        elif "Critical" in entry["Status"]:
            status_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid") # Red
            status_cell.font = Font(color="9C0006")

        ws.cell(row=row_idx, column=3, value=entry["CAS (Inv)"])
        ws.cell(row=row_idx, column=4, value=entry["Physical State"])
        ws.cell(row=row_idx, column=5, value=entry["Source"])
        ws.cell(row=row_idx, column=6, value=entry["Date"])
        ws.cell(row=row_idx, column=7, value=entry["Details (Std)"])

    # Create Table
    ref = f"A1:G{last_row}"
    tab = Table(displayName="AuditTable", ref=ref)
    style = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False,
                           showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    tab.tableStyleInfo = style
    ws.add_table(tab)

    # Auto-width columns
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = min(adjusted_width, 50)

    print(f"Saving Audit Excel to: {output_path}")
    try:
        wb.save(output_path)
        print("Success! Audit file created.")
    except Exception as e:
        print(f"Error saving file: {e}")

if __name__ == "__main__":
    create_audit_excel()
