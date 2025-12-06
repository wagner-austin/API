import polars as pl
from pathlib import Path
import os
from datetime import datetime
from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.table import Table as ExcelTable, TableStyleInfo
from rich.console import Console
from rich.table import Table as RichTable
from rich.panel import Panel

console = Console(force_terminal=True, legacy_windows=False)

def deduplicate_headers(headers):
    """Helper to ensure column names are unique for Polars."""
    counts = {}
    new_headers = []
    for h in headers:
        original = str(h)
        if original in counts:
            counts[original] += 1
            new_headers.append(f"{original}_{counts[original]}")
        else:
            counts[original] = 0
            new_headers.append(original)
    return new_headers


def extract_standards():
    base_path = Path("C:/Users/austi/PROJECTS/UC Irvine/Celia Louise Braun Faiola - FaiolaLab")

    files = {
        "Response Factors": base_path / "Notebooks/Jasmine OseiEnin Lab Notebook/Response factors.xlsx",
        "Soil VOC": base_path / "Current Projects/Soil VOC quantitation.xlsx",
        "Avisa Calc": base_path / "Notebooks/Avisa Lab Notebook/Standard Calculations (1).xlsx",
        "8mix": base_path / "Notebooks/Jasmine OseiEnin Lab Notebook/2023-2024/Summer 24/8mix_calc.xlsx",
        "Std Tidy": base_path / "Notebooks/Jasmine OseiEnin Lab Notebook/2023-2024/Summer 24/std_tidy.xlsx",
        
        # New files identified
        "StandardsAndCals": base_path / "InstrumentLogs/TDGC/Calibrations/StandardsAndCals.xlsx",
        "ChiralStandards": base_path / "InstrumentLogs/TDGC/Calibrations/ChiralStandards_Cal - Updated.xlsx",
        "UniversalList": base_path / "Current Projects/Thermal Stress Project/2021-2022 BVOC collection experiment (Juan)/GCMS data/Universal Chemical List.xlsx",
        "Jasmine2024": base_path / "InstrumentLogs/TDGC/Calibrations/old files/Jasmine Chemcial Standard List 2024.xlsx",
        "ClaireStd": base_path / "InstrumentLogs/TDGC/Calibrations/old files/Claire Chemical Standard List-Faiola.xlsx",
        "OldCompiled": base_path / "InstrumentLogs/TDGC/Calibrations/old files/OLD_CompiledStandardList.xlsx"
    }

    standards_list = []
    seen_names = set()  # Track by normalized name only for deduplication
    file_stats = {}  # Track stats per file

    def add_standard(name, source, date, type_, details):
        import re
        if not name or not isinstance(name, str) or len(name.strip()) < 3:
            return False
        name = name.strip()

        # Strip R-style X prefix from column names (e.g., "X3.Carene" -> "3-Carene")
        if name.startswith("X") and len(name) > 1 and (name[1].isdigit() or name[1] == "."):
            name = name[1:].lstrip(".")

        # Convert dots and underscores to hyphens for consistency
        name = name.replace(".", "-").replace("_", "-")
        # Clean up multiple hyphens
        name = re.sub(r'-+', '-', name)
        name = name.strip("-")

        name_lower = name.lower()

        # Skip obvious non-chemical entries
        skip_exact = [
            "null", "none", "na", "total", "rt", "id", "code", "date", "column", "sheet", "nan",
            "area", "mass", "sample", "control", "chemical name", "compound", "name", "standard",
            "injected volume", "response factor", "relative to", "concentration", "cartridge",
            "point", "volume", "slope", "calc mass", "int area", "peak area", "1ul", "2ul",
            "other", "monoterpene", "monoterpenoid", "sesquiterpene", "alkane",
            "benzoic-acid",  # duplicate with Benzoic acid
            "notes", "achieved", "min area", "max area", "standard ran? (y/n)",
            "compound rf", "ana notes", "claire notes", "reference compound",
            "vial label", "dilute", "concentrate", "mixture"
        ]
        if name_lower in skip_exact:
            return False

        # Skip if starts with these
        skip_startswith = ["sample", "samplw", "tic:", "col-", "column-", "relative to", "standard volume", "unknown", "mt", "omt", "sqt", "osqt"]
        if any(name_lower.startswith(s) for s in skip_startswith):
            return False
            
        # Strict check for placeholder-like headers (e.g. MT1, MT2, OMT1, SQT20) which shouldn't be in standards list
        if re.match(r'^(mt|omt|sqt|osqt|unknown)\d+$', name_lower):
            return False

        # Skip if contains these substrings
        skip_contains = ["\\data-ms", "-d\\", "injected", "response factor", " and u", "(y/n)"]
        if any(s in name_lower for s in skip_contains):
            return False

        # Skip pure numbers (retention times, peak areas, CAS numbers without context)
        if re.match(r'^-?\d+[-\d]*$', name):
            return False

        # Skip if it looks like a formula/equation
        if re.search(r'\*x\s*\+', name):
            return False

        # Skip very long entries (likely descriptions)
        if len(name) > 80:
            return False

        # Normalize the name for deduplication
        norm_name = name_lower
        # Remove all punctuation and spaces for comparison
        norm_name = re.sub(r'[\s\-\,\.\[\]\(\)]+', '', norm_name)
        # Normalize Greek letter prefixes
        norm_name = re.sub(r'^(alpha|a|α)', 'alpha', norm_name)
        norm_name = re.sub(r'^(beta|b|β)', 'beta', norm_name)
        norm_name = re.sub(r'^(gamma|g|y|γ)', 'gamma', norm_name)
        # Remove stereochemistry prefixes for dedup
        norm_name = re.sub(r'^r\+', '', norm_name)  # (R)-(+)
        norm_name = re.sub(r'^s\+', '', norm_name)  # (S)-(+)
        norm_name = re.sub(r'^\+', '', norm_name)  # (+)
        norm_name = re.sub(r'^\-', '', norm_name)  # (-)
        norm_name = re.sub(r'^\+/\-', '', norm_name)  # (+/-)
        norm_name = re.sub(r'^\?', '', norm_name)  # (?)
        norm_name = re.sub(r'^cis', '', norm_name)
        norm_name = re.sub(r'^trans', '', norm_name)

        if norm_name in seen_names:
            return False
        seen_names.add(norm_name)

        # Clean up the display name - prefer Greek letters
        display_name = name
        # Handle "alpha " with space
        if re.match(r'^alpha[\s\-]', display_name, re.I):
            display_name = "α-" + re.sub(r'^alpha[\s\-]+', '', display_name, flags=re.I)
        elif re.match(r'^a-', display_name, re.I):
            display_name = "α-" + display_name[2:]
        elif re.match(r'^beta[\s\-]', display_name, re.I):
            display_name = "β-" + re.sub(r'^beta[\s\-]+', '', display_name, flags=re.I)
        elif re.match(r'^b-', display_name, re.I):
            display_name = "β-" + display_name[2:]
        elif re.match(r'^gamma[\s\-]', display_name, re.I):
            display_name = "γ-" + re.sub(r'^gamma[\s\-]+', '', display_name, flags=re.I)
        elif re.match(r'^y-', display_name, re.I):
            display_name = "γ-" + display_name[2:]

        # Capitalize first letter of chemical name (after Greek prefix)
        if display_name.startswith(("α-", "β-", "γ-")):
            prefix = display_name[:2]
            rest = display_name[2:]
            display_name = prefix + rest[0].upper() + rest[1:] if rest else prefix
        else:
            display_name = display_name[0].upper() + display_name[1:] if display_name else display_name

        standards_list.append({
            "Chemical Name": display_name,
            "Source": source,
            "Date": date,
            "Type": type_,
            "Details": details
        })
        return True

    console.print(Panel.fit("[bold blue]Chemical Standards Extraction[/bold blue]", subtitle="Processing Excel Files"))

    # --- 1. Process Response Factors ---
    console.print(f"\n[bold cyan]1. Response Factors[/bold cyan]: {files['Response Factors'].name}")
    file_stats["Response Factors"] = {"sheets": 0, "extracted": 0}
    try:
        mod_time = os.path.getmtime(files['Response Factors'])
        rf_date = datetime.fromtimestamp(mod_time).strftime('%Y-%m-%d')

        wb_rf = load_workbook(files['Response Factors'], read_only=True)
        rf_sheet_names = wb_rf.sheetnames
        wb_rf.close()
        file_stats["Response Factors"]["sheets"] = len(rf_sheet_names)

        for sheet_name in rf_sheet_names:
            count_before = len(standards_list)
            try:
                df_rf = pl.read_excel(source=files['Response Factors'], sheet_name=sheet_name, engine="openpyxl", has_header=True)

                # Find chemical name column
                chem_col = None
                for col in df_rf.columns:
                    col_lower = col.lower()
                    if "chemical" in col_lower and "name" in col_lower:
                        chem_col = col
                        break
                    elif col_lower == "name" or col_lower == "compound":
                        chem_col = col

                if chem_col:
                    for row in df_rf.iter_rows(named=True):
                        chem = row.get(chem_col)
                        if chem and isinstance(chem, str):
                            density = row.get("Density (g/mL)") or row.get("Density")
                            details = f"Density: {density}" if density else f"Sheet: {sheet_name}"
                            add_standard(chem, "Jasmine - Response Factors", rf_date, "Standard Mix", details)
            except Exception as e:
                console.print(f"    [yellow]Warning[/yellow]: Sheet '{sheet_name}': {e}")

            count_after = len(standards_list)
            extracted = count_after - count_before
            file_stats["Response Factors"]["extracted"] += extracted
            status = "[green]+[/green]" if extracted > 0 else "[dim]-[/dim]"
            console.print(f"    {status} {sheet_name}: [cyan]{extracted}[/cyan] chemicals")
    except Exception as e:
        console.print(f"    [red]Error[/red]: {e}")

    # --- 2. Process Soil VOC Quantitation ---
    console.print(f"\n[bold cyan]2. Soil VOC[/bold cyan]: {files['Soil VOC'].name}")
    file_stats["Soil VOC"] = {"sheets": 0, "extracted": 0}
    try:
        mod_time = os.path.getmtime(files['Soil VOC'])
        soil_file_date = datetime.fromtimestamp(mod_time).strftime('%Y-%m-%d')

        wb_soil = load_workbook(files['Soil VOC'], read_only=True)
        soil_sheet_names = wb_soil.sheetnames
        wb_soil.close()
        file_stats["Soil VOC"]["sheets"] = len(soil_sheet_names)

        for sheet_name in soil_sheet_names:
            count_before = len(standards_list)
            extracted_chems = set()

            try:
                df_raw = pl.read_excel(source=files['Soil VOC'], sheet_name=sheet_name, engine="openpyxl", has_header=False, infer_schema_length=None)
            except Exception as e:
                console.print(f"    [red]x[/red] {sheet_name}: {e}")
                continue

            # Strategy 1: Known structured sheets
            if sheet_name in ["Standard list", "compound_colors (2)", "compound_colors", "Sheet1"]:
                try:
                    df_clean = pl.read_excel(source=files['Soil VOC'], sheet_name=sheet_name, engine="openpyxl", has_header=True)
                    for col in ["name", "compound", "Name", "Compound"]:
                        if col in df_clean.columns:
                            for val in df_clean[col].to_list():
                                if val and isinstance(val, str):
                                    extracted_chems.add(val.strip())
                except:
                    pass

            # Strategy 2: Find header row and extract from columns
            if not extracted_chems and df_raw.height > 0:
                header_row_idx = -1
                for r_idx in range(min(df_raw.height, 5)):
                    row_vals = [str(v).strip().lower() for v in df_raw.row(r_idx) if v is not None]
                    if any(kw in val for val in row_vals for kw in ["compound", "name", "pinene", "terpene", "alkane"]):
                        header_row_idx = r_idx
                        break

                if header_row_idx != -1:
                    headers = [str(val).strip() if val is not None else f"col_{i}" for i, val in enumerate(df_raw.row(header_row_idx))]
                    headers = deduplicate_headers(headers)

                    try:
                        df_processed = pl.DataFrame(df_raw.slice(header_row_idx + 1).rows(), schema=headers, orient="row", infer_schema_length=None)

                        # Extract from column names (chemical names as headers)
                        for col in df_processed.columns:
                            col_lower = col.lower()
                            if any(kw in col_lower for kw in ["pinene", "terpene", "limonene", "alkane", "cyclo"]):
                                extracted_chems.add(col.split("(")[0].strip())

                        # Extract from name/compound columns
                        for col in df_processed.columns:
                            if col.lower() in ["compound", "name", "chemical name", "analyte"]:
                                for val in df_processed[col].to_list():
                                    if val and isinstance(val, str) and len(val.strip()) > 2:
                                        extracted_chems.add(val.strip())
                    except:
                        pass

            for chem in extracted_chems:
                add_standard(chem, "Soil VOC Project", soil_file_date, "Standard", f"Sheet: {sheet_name}")

            count_after = len(standards_list)
            extracted = count_after - count_before
            file_stats["Soil VOC"]["extracted"] += extracted
            status = "[green]+[/green]" if extracted > 0 else "[dim]-[/dim]"
            console.print(f"    {status} {sheet_name}: [cyan]{extracted}[/cyan] chemicals")

    except Exception as e:
        console.print(f"    [red]Error[/red]: {e}")

    # --- 3. Process Avisa Standard Calculations ---
    console.print(f"\n[bold cyan]3. Avisa Calc[/bold cyan]: {files['Avisa Calc'].name}")
    file_stats["Avisa Calc"] = {"sheets": 0, "extracted": 0}
    try:
        mod_time = os.path.getmtime(files['Avisa Calc'])
        avisa_date = datetime.fromtimestamp(mod_time).strftime('%Y-%m-%d')

        wb_avisa = load_workbook(files['Avisa Calc'], read_only=True)
        avisa_sheet_names = wb_avisa.sheetnames
        wb_avisa.close()
        file_stats["Avisa Calc"]["sheets"] = len(avisa_sheet_names)

        for sheet_name in avisa_sheet_names:
            count_before = len(standards_list)

            try:
                df_raw = pl.read_excel(source=files['Avisa Calc'], sheet_name=sheet_name, engine="openpyxl", has_header=False, infer_schema_length=None)
            except Exception as e:
                console.print(f"    [red]x[/red] {sheet_name}: {e}")
                continue

            # Check first cell for chemical name (common pattern in this file)
            if df_raw.height > 0 and df_raw.width > 0:
                first_cell = df_raw[0, 0]
                if isinstance(first_cell, str) and any(kw in first_cell.lower() for kw in ["limonene", "pinene", "camphor", "terpene", "linalool", "eucalyptol"]):
                    add_standard(first_cell.split("(")[0].strip(), "Avisa - Standard Calculations", avisa_date, "Calculated Standard", f"Sheet: {sheet_name}")

            # Find header row
            header_row_idx = -1
            if df_raw.height > 1:
                for r_idx in range(min(df_raw.height, 5)):
                    row_vals = [str(v).strip().lower() for v in df_raw.row(r_idx) if v is not None]
                    if any(kw in val for val in row_vals for kw in ["compound", "name", "standard", "analyte"]):
                        header_row_idx = r_idx
                        break

            if header_row_idx != -1:
                headers = [str(val).strip() if val is not None else f"col_{i}" for i, val in enumerate(df_raw.row(header_row_idx))]
                headers = deduplicate_headers(headers)

                try:
                    df_processed = pl.DataFrame(df_raw.slice(header_row_idx + 1).rows(), schema=headers, orient="row", infer_schema_length=None)

                    for col in df_processed.columns:
                        if any(kw in col.lower() for kw in ["compound", "name", "standard", "analyte"]):
                            for val in df_processed[col].to_list():
                                if val and isinstance(val, str) and len(val.strip()) > 2:
                                    add_standard(val.strip(), "Avisa - Standard Calculations", avisa_date, "Calculated Standard", f"Sheet: {sheet_name}")
                except:
                    pass

            count_after = len(standards_list)
            extracted = count_after - count_before
            file_stats["Avisa Calc"]["extracted"] += extracted
            status = "[green]+[/green]" if extracted > 0 else "[dim]-[/dim]"
            console.print(f"    {status} {sheet_name}: [cyan]{extracted}[/cyan] chemicals")

    except Exception as e:
        console.print(f"    [red]Error[/red]: {e}")

    # --- 4. 8mix_calc ---
    console.print(f"\n[bold cyan]4. 8mix[/bold cyan]: {files['8mix'].name}")
    file_stats["8mix"] = {"sheets": 0, "extracted": 0}
    try:
        mod_time = os.path.getmtime(files['8mix'])
        mix_date = datetime.fromtimestamp(mod_time).strftime('%Y-%m-%d')

        wb_8mix = load_workbook(files['8mix'], read_only=True)
        mix_sheet_names = wb_8mix.sheetnames
        wb_8mix.close()
        file_stats["8mix"]["sheets"] = len(mix_sheet_names)

        for sheet_name in mix_sheet_names:
            count_before = len(standards_list)

            try:
                df_raw = pl.read_excel(source=files['8mix'], sheet_name=sheet_name, engine="openpyxl", has_header=False, infer_schema_length=None)
            except:
                continue

            # Find header row with "concentration"
            header_row_idx = -1
            if df_raw.height > 0:
                for r_idx in range(min(df_raw.height, 5)):
                    row_vals = [str(v).strip().lower() for v in df_raw.row(r_idx) if v is not None]
                    if "concentration" in row_vals:
                        header_row_idx = r_idx
                        break

            if header_row_idx != -1:
                headers = [str(val).strip() if val is not None else f"col_{i}" for i, val in enumerate(df_raw.row(header_row_idx))]
                headers = deduplicate_headers(headers)

                # Extract chemical names from column headers
                skip_cols = ["concentration", "standard", "cartridge", "slope", "rt", "calc mass", "column", ""]
                for col in headers:
                    col_lower = col.lower()
                    if not any(skip in col_lower for skip in skip_cols) and not col_lower.endswith("_1") and len(col.strip()) > 2:
                        add_standard(col.split("(")[0].strip(), "Avisa - 8mix", mix_date, "8-Mix Component", f"Sheet: {sheet_name}")
            else:
                # Scan for known chemical names
                for r_idx in range(min(df_raw.height, 5)):
                    for val in df_raw.row(r_idx):
                        if isinstance(val, str) and any(kw in val.lower() for kw in ["pinene", "terpene", "limonene", "thujone", "linalool", "eucalyptol", "myrcene"]):
                            add_standard(val.split("(")[0].strip(), "Avisa - 8mix", mix_date, "8-Mix Component", f"Sheet: {sheet_name}")

            count_after = len(standards_list)
            extracted = count_after - count_before
            file_stats["8mix"]["extracted"] += extracted
            status = "[green]+[/green]" if extracted > 0 else "[dim]-[/dim]"
            console.print(f"    {status} {sheet_name}: [cyan]{extracted}[/cyan] chemicals")

    except Exception as e:
        console.print(f"    [red]Error[/red]: {e}")

    # --- 5. Std Tidy ---
    console.print(f"\n[bold cyan]5. Std Tidy[/bold cyan]: {files['Std Tidy'].name}")
    file_stats["Std Tidy"] = {"sheets": 0, "extracted": 0}
    try:
        mod_time = os.path.getmtime(files['Std Tidy'])
        tidy_date = datetime.fromtimestamp(mod_time).strftime('%Y-%m-%d')

        wb_tidy = load_workbook(files['Std Tidy'], read_only=True)
        tidy_sheet_names = wb_tidy.sheetnames
        wb_tidy.close()
        file_stats["Std Tidy"]["sheets"] = len(tidy_sheet_names)

        for sheet_name in tidy_sheet_names:
            count_before = len(standards_list)
            extracted_chems = set()

            try:
                df_raw = pl.read_excel(source=files['Std Tidy'], sheet_name=sheet_name, engine="openpyxl", has_header=False, infer_schema_length=None)
            except:
                continue

            # Check for chemical.name column (Sheet2 pattern)
            try:
                df_with_header = pl.read_excel(source=files['Std Tidy'], sheet_name=sheet_name, engine="openpyxl", has_header=True)
                for col in df_with_header.columns:
                    if "chemical" in col.lower() and "name" in col.lower():
                        for val in df_with_header[col].to_list():
                            if val and isinstance(val, str):
                                # Clean up names like "alpha.pinene" -> "alpha-pinene"
                                clean_name = val.replace(".", "-").strip()
                                extracted_chems.add(clean_name)
            except:
                pass

            # Extract from column headers (Sheet1 pattern - chemicals like "A-pinene (5.4 min) Int area")
            if df_raw.height > 0:
                header_row = df_raw.row(0)
                for val in header_row:
                    if val and isinstance(val, str):
                        val_lower = val.lower()
                        if any(kw in val_lower for kw in ["pinene", "terpene", "myrcene", "linalool", "eucalyptol", "thujone"]):
                            # Extract just the chemical name part
                            name = val.split("(")[0].split("Int")[0].split("mass")[0].strip()
                            if len(name) > 2:
                                extracted_chems.add(name)

            for chem in extracted_chems:
                add_standard(chem, "Avisa - Tidy Standards", tidy_date, "Standard Mix Component", f"Sheet: {sheet_name}")

            count_after = len(standards_list)
            extracted = count_after - count_before
            file_stats["Std Tidy"]["extracted"] += extracted
            status = "[green]+[/green]" if extracted > 0 else "[dim]-[/dim]"
            console.print(f"    {status} {sheet_name}: [cyan]{extracted}[/cyan] chemicals")

    except Exception as e:
        console.print(f"    [red]Error[/red]: {e}")

    # --- 6. StandardsAndCals ---
    console.print(f"\n[bold cyan]6. StandardsAndCals[/bold cyan]: {files['StandardsAndCals'].name}")
    file_stats["StandardsAndCals"] = {"sheets": 0, "extracted": 0}
    try:
        mod_time = os.path.getmtime(files['StandardsAndCals'])
        sc_date = datetime.fromtimestamp(mod_time).strftime('%Y-%m-%d')
        
        # Explicitly processing 'Work list' as it has the mix data
        try:
            df_wl = pl.read_excel(source=files['StandardsAndCals'], sheet_name="Work list", engine="openpyxl", has_header=True)
            
            # Find the mixture column
            mix_col = None
            for col in df_wl.columns:
                if "mixture" in col.lower() or "arrangment" in col.lower(): # Note: Typo 'arrangment' in source
                    mix_col = col
                    break
            
            if mix_col:
                count_before = len(standards_list)
                for val in df_wl[mix_col].to_list():
                    if val and isinstance(val, str):
                        # Split mixtures separated by '/'
                        parts = val.split("/")
                        for part in parts:
                            # Clean individual chemical name
                            cleaned = part.strip()
                            add_standard(cleaned, "StandardsAndCals - Work list", sc_date, "Mix Component", f"From mix: {val[:30]}...")
                
                extracted = len(standards_list) - count_before
                file_stats["StandardsAndCals"]["extracted"] += extracted
                console.print(f"    [green]+[/green] Work list: [cyan]{extracted}[/cyan] chemicals")
                file_stats["StandardsAndCals"]["sheets"] = 1 # Approximate since we focused on 1 sheet
        except Exception as e:
             console.print(f"    [red]Error[/red] reading Work list: {e}")
    except Exception as e:
        console.print(f"    [red]Error[/red]: {e}")


    # --- 7. ChiralStandards ---
    console.print(f"\n[bold cyan]7. ChiralStandards[/bold cyan]: {files['ChiralStandards'].name}")
    file_stats["ChiralStandards"] = {"sheets": 0, "extracted": 0}
    try:
        mod_time = os.path.getmtime(files['ChiralStandards'])
        chiral_date = datetime.fromtimestamp(mod_time).strftime('%Y-%m-%d')
        
        # Retention Times Sheet
        try:
            df_rt = pl.read_excel(source=files['ChiralStandards'], sheet_name="Retention Times", engine="openpyxl", has_header=True)
            count_before = len(standards_list)
            if "Compound" in df_rt.columns:
                for val in df_rt["Compound"].to_list():
                    if val: add_standard(val, "ChiralStandards - RT", chiral_date, "Chiral Standard", "Retention Times Sheet")
            
            extracted = len(standards_list) - count_before
            file_stats["ChiralStandards"]["extracted"] += extracted
            console.print(f"    [green]+[/green] Retention Times: [cyan]{extracted}[/cyan] chemicals")
            file_stats["ChiralStandards"]["sheets"] = 1
        except: pass

    except Exception as e:
        console.print(f"    [red]Error[/red]: {e}")


    # --- 8. UniversalList ---
    console.print(f"\n[bold cyan]8. UniversalList[/bold cyan]: {files['UniversalList'].name}")
    file_stats["UniversalList"] = {"sheets": 0, "extracted": 0}
    try:
        mod_time = os.path.getmtime(files['UniversalList'])
        univ_date = datetime.fromtimestamp(mod_time).strftime('%Y-%m-%d')

        # Standards list sheet
        try:
            df_std = pl.read_excel(source=files['UniversalList'], sheet_name="Standards list", engine="openpyxl", has_header=True)
            count_before = len(standards_list)
            # Column is "Chemical " with space
            col_name = next((c for c in df_std.columns if "chemical" in c.lower()), None)
            if col_name:
                for val in df_std[col_name].to_list():
                    if val: add_standard(val, "UniversalList - Standards", univ_date, "Standard", "Standards list sheet")
            
            extracted = len(standards_list) - count_before
            file_stats["UniversalList"]["extracted"] += extracted
            console.print(f"    [green]+[/green] Standards list: [cyan]{extracted}[/cyan] chemicals")
        except: pass
        
        # RT combined Sheet - Extract Headers
        try:
            df_rt = pl.read_excel(source=files['UniversalList'], sheet_name="RT combined(in progress)", engine="openpyxl", has_header=True)
            count_before = len(standards_list)
            for col in df_rt.columns:
                # Strict validation handled by add_standard (excludes MT#, SQT#, etc)
                add_standard(col, "UniversalList - RT Combined", univ_date, "Tracked Compound", "Column Header")
            
            extracted = len(standards_list) - count_before
            file_stats["UniversalList"]["extracted"] += extracted
            console.print(f"    [green]+[/green] RT combined: [cyan]{extracted}[/cyan] chemicals")
            file_stats["UniversalList"]["sheets"] = 2
        except: pass

    except Exception as e:
        console.print(f"    [red]Error[/red]: {e}")

    # --- 9. Jasmine2024 ---
    console.print(f"\n[bold cyan]9. Jasmine2024[/bold cyan]: {files['Jasmine2024'].name}")
    file_stats["Jasmine2024"] = {"sheets": 0, "extracted": 0}
    try:
        mod_time = os.path.getmtime(files['Jasmine2024'])
        jas_date = datetime.fromtimestamp(mod_time).strftime('%Y-%m-%d')
        
        df_j = pl.read_excel(source=files['Jasmine2024'], sheet_name="Sheet1", engine="openpyxl", has_header=True)
        count_before = len(standards_list)
        col_name = next((c for c in df_j.columns if "chemical" in c.lower()), None)
        if col_name:
            for val in df_j[col_name].to_list():
                if val: add_standard(val, "Jasmine 2024 List", jas_date, "Standard", "Sheet1")
        
        extracted = len(standards_list) - count_before
        file_stats["Jasmine2024"]["extracted"] += extracted
        console.print(f"    [green]+[/green] Sheet1: [cyan]{extracted}[/cyan] chemicals")
        file_stats["Jasmine2024"]["sheets"] = 1
    except Exception as e:
        console.print(f"    [red]Error[/red]: {e}")

    # --- 10. ClaireStd ---
    console.print(f"\n[bold cyan]10. ClaireStd[/bold cyan]: {files['ClaireStd'].name}")
    file_stats["ClaireStd"] = {"sheets": 0, "extracted": 0}
    try:
        mod_time = os.path.getmtime(files['ClaireStd'])
        claire_date = datetime.fromtimestamp(mod_time).strftime('%Y-%m-%d')
        
        df_c = pl.read_excel(source=files['ClaireStd'], sheet_name="Sheet1", engine="openpyxl", has_header=True)
        count_before = len(standards_list)
        if "Compound" in df_c.columns:
            for val in df_c["Compound"].to_list():
                if val: add_standard(val, "Claire Faiola List", claire_date, "Standard", "Sheet1")
        
        extracted = len(standards_list) - count_before
        file_stats["ClaireStd"]["extracted"] += extracted
        console.print(f"    [green]+[/green] Sheet1: [cyan]{extracted}[/cyan] chemicals")
        file_stats["ClaireStd"]["sheets"] = 1
    except Exception as e:
        console.print(f"    [red]Error[/red]: {e}")
        
    # --- 11. OldCompiled ---
    console.print(f"\n[bold cyan]11. OldCompiled[/bold cyan]: {files['OldCompiled'].name}")
    file_stats["OldCompiled"] = {"sheets": 0, "extracted": 0}
    try:
        mod_time = os.path.getmtime(files['OldCompiled'])
        old_date = datetime.fromtimestamp(mod_time).strftime('%Y-%m-%d')
        
        # Sheet: Rearrangment
        try:
            df_old = pl.read_excel(source=files['OldCompiled'], sheet_name="Rearrangment", engine="openpyxl", has_header=True)
            count_before = len(standards_list)
            col_name = "Compiled standard list"
            if col_name in df_old.columns:
                for val in df_old[col_name].to_list():
                    if val: add_standard(val, "Old Compiled List", old_date, "Historical Standard", "Rearrangment Sheet")
            
            extracted = len(standards_list) - count_before
            file_stats["OldCompiled"]["extracted"] += extracted
            console.print(f"    [green]+[/green] Rearrangment: [cyan]{extracted}[/cyan] chemicals")
            file_stats["OldCompiled"]["sheets"] = 1
        except: pass
    except Exception as e:
        console.print(f"    [red]Error[/red]: {e}")


    # --- Summary Table ---
    console.print()
    summary_table = RichTable(title="[bold]Extraction Summary[/bold]")
    summary_table.add_column("File", style="cyan")
    summary_table.add_column("Sheets", justify="right")
    summary_table.add_column("Extracted", justify="right", style="green")

    total_sheets = 0
    total_extracted = 0
    for file_name, stats in file_stats.items():
        summary_table.add_row(file_name, str(stats["sheets"]), str(stats["extracted"]))
        total_sheets += stats["sheets"]
        total_extracted += stats["extracted"]

    summary_table.add_row("[bold]TOTAL[/bold]", f"[bold]{total_sheets}[/bold]", f"[bold]{len(standards_list)}[/bold]")
    console.print(summary_table)

    # --- Save to Excel ---
    output_path = base_path / "Notebooks/Emily Truong Notebook/Chemical_Standards_List_2025.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "Chemical Standards"

    headers = ["Chemical Name", "Source", "Date", "Type", "Details"]
    for col_idx, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_idx, value=header)

    for row_idx, entry in enumerate(standards_list, 2):
        ws.cell(row=row_idx, column=1, value=entry["Chemical Name"])
        ws.cell(row=row_idx, column=2, value=entry["Source"])
        ws.cell(row=row_idx, column=3, value=entry["Date"])
        ws.cell(row=row_idx, column=4, value=entry["Type"])
        ws.cell(row=row_idx, column=5, value=entry["Details"])

    last_row = len(standards_list) + 1
    if last_row > 1:
        tab = ExcelTable(displayName="ChemicalStandards2025", ref=f"A1:E{last_row}")
        tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
        ws.add_table(tab)

    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 60)

    wb.save(output_path)
    console.print(f"\n[bold green]Saved {len(standards_list)} unique standards[/bold green] to:")
    console.print(f"  [dim]{output_path}[/dim]")


if __name__ == "__main__":
    extract_standards()
