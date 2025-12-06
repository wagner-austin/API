
import polars as pl
from pathlib import Path
from rich.console import Console
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter

def create_2021_inventory_report():
    console = Console()
    base_path = Path(r"C:\Users\austi\PROJECTS\UC Irvine\Celia Louise Braun Faiola - FaiolaLab")
    
    # Input File
    input_path = base_path / "Important Docs/Chemical Inventory/02252021-Chemical Inventory.xlsx"
    
    # Output File
    output_path = base_path / "Notebooks/Emily Truong Notebook/Chemical_Inventory_List_2021.xlsx"

    console.print(f"[bold blue]--- Processing 2021 Inventory ---[/bold blue]")
    console.print(f"Input: {input_path}")
    
    combined_data = []

    # --- Process CiBR-Trac Sheet ---
    try:
        console.print("Reading 'CiBR-Trac' sheet...")
        df_cibr = pl.read_excel(source=input_path, sheet_name="CiBR-Trac", engine="openpyxl")
        
        # Normalize CiBR-Trac
        # Expected columns: Chemical_Name, CAS, Chemical_Physical_State, Container_Size, Units, Container_Number
        # We want: Chemical Name, CAS, Physical State, Amount, Units, Containers, Location, Source
        
        df_cibr = df_cibr.select([
            pl.col("Chemical_Name").alias("Chemical Name"),
            pl.col("CAS"),
            pl.col("Chemical_Physical_State").alias("Physical State"),
            pl.col("Container_Size").cast(pl.Float64, strict=False).alias("Amount"),
            pl.col("Units"),
            pl.col("Container_Number").cast(pl.Int64, strict=False).alias("Containers"),
            pl.lit("CiBR-Trac").alias("Location"),
            pl.lit("2021 Inventory (CiBR-Trac)").alias("Source")
        ])
        
        combined_data.append(df_cibr)
        console.print(f"  Loaded {len(df_cibr)} rows.")
        
    except Exception as e:
        console.print(f"[red]Error reading CiBR-Trac: {e}[/red]")

    # --- Process Room 428 Sheet ---
    try:
        console.print("Reading '428' sheet...")
        df_428 = pl.read_excel(source=input_path, sheet_name="428", engine="openpyxl")
        
        # Normalize 428
        # Expected columns: Chemical Name, CAS, Physical State, Amount, Units, Room Number, Sublocation
        
        # Construct Location from Room Number and Sublocation
        df_428 = df_428.with_columns(
            pl.concat_str([
                pl.lit("Room "),
                pl.col("Room Number").cast(pl.Utf8),
                pl.lit(" - "),
                pl.col("Sublocation")
            ]).alias("Location")
        )

        # 428 doesn't have a "Containers" count column, assume 1 per row? Or just leave null.
        # Let's set Containers to 1 as a default if missing, or null.
        # The other script set it to 1.
        
        df_428 = df_428.select([
            pl.col("Chemical Name"),
            pl.col("CAS"),
            pl.col("Physical State"),
            pl.col("Amount").cast(pl.Float64, strict=False),
            pl.col("Units"),
            pl.lit(1).cast(pl.Int64).alias("Containers"), # Assumption
            pl.col("Location"),
            pl.lit("2021 Inventory (Room 428)").alias("Source")
        ])
        
        combined_data.append(df_428)
        console.print(f"  Loaded {len(df_428)} rows.")

    except Exception as e:
        console.print(f"[red]Error reading 428: {e}[/red]")

    if not combined_data:
        console.print("[red]No data loaded. Exiting.[/red]")
        return

    # Combine DataFrames
    final_df = pl.concat(combined_data)
    
    # Cast types for consistency (Handle mixed types in CAS or Amount)
    final_df = final_df.with_columns([
        pl.col("Chemical Name").cast(pl.Utf8),
        pl.col("CAS").cast(pl.Utf8),
        pl.col("Physical State").cast(pl.Utf8),
        pl.col("Amount").cast(pl.Float64, strict=False),
        pl.col("Units").cast(pl.Utf8),
        pl.col("Containers").cast(pl.Int64, strict=False),
        pl.col("Location").cast(pl.Utf8),
        pl.col("Source").cast(pl.Utf8)
    ])

    console.print(f"Total Combined Rows: {len(final_df)}")

    # --- Write to Excel with Formatting ---
    console.print(f"Writing to: {output_path}")
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Chemical_List_2021"

    # Write Headers
    headers = final_df.columns
    for col_idx, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_idx, value=header)

    # Write Data
    for row_idx, row in enumerate(final_df.iter_rows(), 2):
        for col_idx, value in enumerate(row, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    # Create Excel Table
    last_row = len(final_df) + 1
    last_col_letter = get_column_letter(len(headers))
    table_ref = f"A1:{last_col_letter}{last_row}"
    
    tab = Table(displayName="Inventory2021", ref=table_ref)
    
    # Style: TableStyleMedium9 (Blue, with headers, banded rows)
    style = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False
    )
    tab.tableStyleInfo = style
    ws.add_table(tab)

    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        # Cap width at 60
        ws.column_dimensions[column].width = min(adjusted_width, 60)

    # Save
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    console.print(f"[bold green]Successfully created: {output_path}[/bold green]")

if __name__ == "__main__":
    create_2021_inventory_report()
