
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo

def update_inventory_metadata():
    # Define path
    excel_path = Path(r"C:\Users\austi\PROJECTS\UC Irvine\Celia Louise Braun Faiola - FaiolaLab\Notebooks\Emily Truong Notebook\Chemical_Inventory_List_2025.xlsx")

    print(f"Updating Excel file: {excel_path}")
    
    try:
        wb = load_workbook(excel_path)
        ws = wb.active
    except FileNotFoundError:
        print("Error: Excel file not found.")
        return
    except Exception as e:
        print(f"Error reading Excel: {e}")
        return

    headers = [cell.value for cell in ws[1]]
    max_row = ws.max_row

    # --- Update Source Column ---
    if "Source" in headers:
        print("Source column already exists. Updating values...")
        source_col_idx = headers.index("Source") + 1
    else:
        print("Adding Source column...")
        source_col_idx = len(headers) + 1
        ws.cell(row=1, column=source_col_idx, value="Source")
        headers.append("Source") # Keep headers list in sync

    for row in range(2, max_row + 1):
        ws.cell(row=row, column=source_col_idx, value="UCI Chemical Inventory (Risk & Safety)")
    
    # --- Update Date Column ---
    if "Date" in headers:
        print("Date column already exists. Updating values...")
        date_col_idx = headers.index("Date") + 1
    else:
        print("Adding Date column...")
        date_col_idx = len(headers) + 1
        ws.cell(row=1, column=date_col_idx, value="Date")
        headers.append("Date")

    for row in range(2, max_row + 1):
        ws.cell(row=row, column=date_col_idx, value="2025-12-05")


    # Update Table Range
    if ws.tables:
        table_name = list(ws.tables.keys())[0]
        tab = ws.tables[table_name]
        
        from openpyxl.utils import get_column_letter
        # The last column is now the max of whatever we added
        last_col_idx = len(headers)
        last_col_letter = get_column_letter(last_col_idx)
        new_ref = f"A1:{last_col_letter}{max_row}"
        
        tab.ref = new_ref
        print(f"Updated Table '{table_name}' range to {new_ref}")

    # Auto-adjust column widths
    ws.column_dimensions[get_column_letter(source_col_idx)].width = 40
    ws.column_dimensions[get_column_letter(date_col_idx)].width = 15

    # Save
    print(f"Saving updated Excel file...")
    try:
        wb.save(excel_path)
        print("Success! Inventory file updated with Source and Date.")
    except Exception as e:
        print(f"Error saving file (is it open?): {e}")

if __name__ == "__main__":
    update_inventory_metadata()
