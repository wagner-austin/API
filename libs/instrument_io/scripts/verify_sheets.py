
from pathlib import Path
from openpyxl import load_workbook
from rich.console import Console
from rich.table import Table

def verify_all_sheets():
    base_path = Path(r"C:\Users\austi\PROJECTS\UC Irvine\Celia Louise Braun Faiola - FaiolaLab")
    
    files_to_check = [
        base_path / "Notebooks/Jasmine OseiEnin Lab Notebook/Response factors.xlsx",
        base_path / "Current Projects/Soil VOC quantitation.xlsx",
        base_path / "Notebooks/Avisa Lab Notebook/Standard Calculations (1).xlsx",
        base_path / "Notebooks/Jasmine OseiEnin Lab Notebook/2023-2024/Summer 24/8mix_calc.xlsx",
        base_path / "Notebooks/Jasmine OseiEnin Lab Notebook/2023-2024/Summer 24/std_tidy.xlsx",
        base_path / "Important Docs/Chemical Inventory/02252021-Chemical Inventory.xlsx"
    ]

    console = Console()
    table = Table(title="[bold blue]Sheet Verification Audit[/bold blue]")
    table.add_column("File Name", style="cyan")
    table.add_column("Total Sheets", style="magenta")
    table.add_column("Sheet Names", style="green")

    print("Scanning files for hidden sheets...")

    for file_path in files_to_check:
        try:
            # read_only=True is faster and safer for just checking names
            wb = load_workbook(file_path, read_only=True)
            sheet_names = wb.sheetnames
            
            table.add_row(
                file_path.name,
                str(len(sheet_names)),
                ", ".join(sheet_names)
            )
            wb.close()
        except Exception as e:
            table.add_row(
                file_path.name,
                "ERROR",
                str(e)
            )

    console.print(table)

if __name__ == "__main__":
    verify_all_sheets()
