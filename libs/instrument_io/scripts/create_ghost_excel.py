
import polars as pl
from pathlib import Path
from rich.console import Console
from rich.table import Table
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import PatternFill, Font, Alignment

def create_formatted_ghost_excel():
    base_path = Path(r"C:\Users\austi\PROJECTS\UC Irvine\Celia Louise Braun Faiola - FaiolaLab")
    
    # 2021 Inventory - This was initially an Excel file, the doc mentioned it was structured.
    inv_2021_path = base_path / "Important Docs/Chemical Inventory/02252021-Chemical Inventory.xlsx"
    
    # 2025 Inventory (the current one)
    inv_2025_path = base_path / "Notebooks/Emily Truong Notebook/Chemical_Inventory_List_2025.xlsx"
    
    output_excel_path = base_path / "Notebooks/Emily Truong Notebook/Chemical_Ghost_Inventory_Report_2025.xlsx"

    console = Console()
    console.print(f"[bold blue]--- Generating Formatted Ghost Inventory Report ---[/bold blue]")

    # --- Load 2021 Inventory ---
    df_2021 = None
    try:
        # Explicitly try to read known sheet names since auto-discovery failed
        known_sheets = ["CiBR-Trac", "428"]
        df_2021_list = []
        
        for sheet in known_sheets:
            try:
                df_sheet = pl.read_excel(source=inv_2021_path, sheet_name=sheet, engine="openpyxl")
                # Add source column
                df_sheet = df_sheet.with_columns(pl.lit(sheet).alias("Original Sheet"))
                df_2021_list.append(df_sheet)
                console.print(f"Successfully read sheet '{sheet}'")
            except Exception as e:
                console.print(f"Could not read sheet '{sheet}': {e}")

        if df_2021_list:
            # For concatenation, we need matching columns. These sheets have different schemas.
            # Strategy: Normalize column names (map to "Chemical Name") and select common/relevant cols.
            
            normalized_dfs = []
            for df in df_2021_list:
                # Check for Chemical Name column
                cols = df.columns
                # 2021 CiBR-Trac has "Chemical_Name", 2021 428 has "Chemical Name"
                target_col = "Chemical_Name" # We will standardize to this
                
                if "Chemical_Name" in cols:
                    pass # Good
                elif "Chemical Name" in cols:
                    df = df.rename({"Chemical Name": "Chemical_Name"})
                
                # Handle CAS
                if "CAS" not in df.columns:
                    df = df.with_columns(pl.lit(None).alias("CAS"))
                
                # Handle Physical State
                # CiBR-Trac: "Chemical_Physical_State"
                # 428: "Physical State"
                if "Chemical_Physical_State" in df.columns:
                    pass
                elif "Physical State" in df.columns:
                    df = df.rename({"Physical State": "Chemical_Physical_State"})
                else:
                    df = df.with_columns(pl.lit(None).alias("Chemical_Physical_State"))

                # Select only key columns to ensure schema compatibility for concat
                selected_df = df.select([
                    pl.col("Chemical_Name").cast(pl.Utf8),
                    pl.col("CAS").cast(pl.Utf8),
                    pl.col("Chemical_Physical_State").cast(pl.Utf8),
                    pl.col("Original Sheet").cast(pl.Utf8)
                ])
                normalized_dfs.append(selected_df)

            df_2021 = pl.concat(normalized_dfs, how="vertical").unique(subset=["Chemical_Name", "CAS"])
            console.print(f"Loaded 2021 Inventory (combined): {len(df_2021)} unique entries")
            
        else:
            # Fallback to default single sheet read if explicit names fail
            console.print("Explicit sheet read failed. Trying default...")
            df_2021 = pl.read_excel(source=inv_2021_path, engine="openpyxl")
            df_2021 = df_2021.with_columns(pl.lit("Unknown Sheet").alias("Original Sheet"))

    except Exception as e:
        console.print(f"[red]Error loading 2021 Inventory '{inv_2021_path}': {e}[/red]")
        return
    
    # --- Load 2025 Inventory ---
    df_2025 = None
    try:
        df_2025 = pl.read_excel(source=inv_2025_path, engine="openpyxl")
    except Exception as e:
        console.print(f"[red]Error loading 2025 Inventory '{inv_2025_path}': {e}[/red]")
        return

    if df_2021 is None or df_2025 is None:
        console.print("[red]Cannot generate report: one or both inventories failed to load.[/red]")
        return

    # --- Normalize Names ---
    def normalize_name_polars(df, col_name, new_col_name):
        if col_name in df.columns:
            return df.with_columns(
                pl.col(col_name).cast(pl.Utf8).str.strip_chars().str.to_lowercase().alias(new_col_name)
            )
        return df # Return original if column not found

    df_2021 = normalize_name_polars(df_2021, "Chemical_Name", "Chemical_Name_normalized")
    df_2025 = normalize_name_polars(df_2025, "Chemical Name", "Chemical Name_normalized")

    df_2021_valid = df_2021.filter(pl.col("Chemical_Name_normalized").is_not_null() & (pl.col("Chemical_Name_normalized") != ""))
    df_2025_valid = df_2025.filter(pl.col("Chemical Name_normalized").is_not_null() & (pl.col("Chemical Name_normalized") != ""))

    names_2021 = set(df_2021_valid["Chemical_Name_normalized"].to_list())
    names_2025 = set(df_2025_valid["Chemical Name_normalized"].to_list())

    # --- Find Ghost Inventory (in 2021 but NOT in 2025) ---
    ghost_chemicals_normalized = names_2021 - names_2025

    if not ghost_chemicals_normalized:
        console.print("[green]Good news! No 'ghost' chemicals found from 2021 inventory.[/green]")
        return

    ghost_data_df = df_2021_valid.filter(pl.col("Chemical_Name_normalized").is_in(list(ghost_chemicals_normalized)))

    # --- Create Formatted Excel Report ---
    wb = Workbook()
    ws = wb.active
    ws.title = "Ghost Inventory Report"

    # Headers for the Excel (from the ghost_data_df DataFrame)
    excel_headers = ["Chemical Name (2021)", "CAS (2021)", "Chemical Physical State (2021)", "Original Sheet (2021)", "Source (2021)"]
    
    # Map DataFrame columns to desired Excel headers
    df_cols_to_extract = ["Chemical_Name", "CAS", "Chemical_Physical_State", "Original Sheet"]
    
    # Add a fixed source for 2021 entries
    ghost_data_df = ghost_data_df.with_columns(pl.lit("UCI Chemical Inventory (2021)").alias("Source (2021)"))
    df_cols_to_extract.append("Source (2021)")


    for col_idx, header in enumerate(excel_headers, 1):
        ws.cell(row=1, column=col_idx, value=header)

    # Write Data
    last_row = 1
    for row_idx, row_data in enumerate(ghost_data_df.iter_rows(named=True), 2):
        last_row = row_idx
        ws.cell(row=row_idx, column=1, value=row_data.get("Chemical_Name"))
        ws.cell(row=row_idx, column=2, value=str(row_data.get("CAS", "")))
        ws.cell(row=row_idx, column=3, value=row_data.get("Chemical_Physical_State", ""))
        ws.cell(row=row_idx, column=4, value=row_data.get("Original Sheet", ""))
        ws.cell(row=row_idx, column=5, value=row_data.get("Source (2021)", ""))


    # Create an Excel Table
    if last_row > 1: # Only create table if there is data
        last_col_letter = 'E' # Assuming 5 columns for now
        ref = f"A1:{last_col_letter}{last_row}"
        tab = Table(displayName="GhostInventory", ref=ref)
        style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                               showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        tab.tableStyleInfo = style
        ws.add_table(tab)

    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = min(adjusted_width, 60)

    # Save the workbook
    try:
        wb.save(output_excel_path)
        console.print(f"\nFormatted Ghost Inventory Report saved to: {output_excel_path}")
    except Exception as e:
        console.print(f"[red]Error saving formatted Excel report: {e}[/red]")

if __name__ == "__main__":
    create_formatted_ghost_excel()
