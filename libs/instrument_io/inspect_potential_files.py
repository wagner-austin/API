
import polars as pl
from pathlib import Path
from rich.console import Console

def inspect_files():
    console = Console()
    base_path = Path(r"C:\Users\austi\PROJECTS\UC Irvine\Celia Louise Braun Faiola - FaiolaLab")
    
    files_to_inspect = [
        "InstrumentLogs/TDGC/Calibrations/StandardsAndCals.xlsx",
        "InstrumentLogs/TDGC/Calibrations/ChiralStandards_Cal - Updated.xlsx",
        "Current Projects/Thermal Stress Project/2021-2022 BVOC collection experiment (Juan)/GCMS data/Universal Chemical List.xlsx",
        "InstrumentLogs/TDGC/Calibrations/old files/Jasmine Chemcial Standard List 2024.xlsx",
        "InstrumentLogs/TDGC/Calibrations/old files/Claire Chemical Standard List-Faiola.xlsx",
        "InstrumentLogs/TDGC/Calibrations/old files/OLD_CompiledStandardList.xlsx"
    ]

    for file_rel in files_to_inspect:
        file_path = base_path / file_rel
        console.print(f"\n[bold blue]Inspecting: {file_path.name}[/bold blue]")
        
        if not file_path.exists():
            console.print("[red]File not found[/red]")
            continue

        try:
            # Get sheet names first
            # Polars read_excel can read specific sheets, but to list them we might need openpyxl directly 
            # or just try reading default and see what happens, but openpyxl is safer for inspection.
            import openpyxl
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            sheets = wb.sheetnames
            console.print(f"Sheets: {sheets}")
            wb.close()

            # Read first few rows of first sheet (or interesting sheets)
            for sheet in sheets[:3]: # Limit to first 3 sheets to save space
                console.print(f"  [cyan]Sheet: {sheet}[/cyan]")
                try:
                    df = pl.read_excel(source=file_path, sheet_name=sheet, engine="openpyxl", has_header=True)
                    if df.height > 0:
                        console.print(f"    Columns: {df.columns}")
                        # Print first 2 rows
                        for row in df.head(2).iter_rows():
                            console.print(f"    Row: {row}")
                    else:
                        console.print("    (Empty sheet)")
                except Exception as e:
                    console.print(f"    [yellow]Could not read with headers: {e}[/yellow]")
                    # Try without headers
                    try:
                        df = pl.read_excel(source=file_path, sheet_name=sheet, engine="openpyxl", has_header=False)
                        console.print(f"    Row 0 (raw): {df.row(0)}")
                    except:
                        pass
        except Exception as e:
            console.print(f"[red]Error inspecting file: {e}[/red]")

if __name__ == "__main__":
    inspect_files()
