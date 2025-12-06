
import sys
from pathlib import Path
from rich.console import Console

# Add the instrument_io library to the python path
sys.path.append(str(Path(__file__).parent / 'src'))

from instrument_io.readers.docx import DOCXReader
from instrument_io.readers.excel import ExcelReader
from instrument_io.readers.pdf import PDFReader # Assuming PDFReader exists and works

def summarize_documents_at_once():
    console = Console()
    base_notebooks_path = Path(r"C:\Users\austi\PROJECTS\UC Irvine\Celia Louise Braun Faiola - FaiolaLab\Notebooks")
    
    files_to_inspect_rel_paths = [
        "Jaycee Fahrner Lab Notebook/SOPs and Manuals/2022 Root_Rhizosphere Metabolomics SOP_.docx",
        "Jaycee Fahrner Lab Notebook/SOPs and Manuals/Copy of 2022 Plant Foliar Metabolomics SOP_.docx",
        "Jaycee Fahrner Lab Notebook/SOPs and Manuals/Plant Foliar Metabolome Method For Monday_.docx",
        "Juan Flores Lab Notebook/BVOC_Sample_Collection_Protocol.docx",
        "Jasmine OseiEnin Lab Notebook/GCMS_Run_2024.xlsx",
        "Maria Flores Lab Notebook/SMEAR II Hyytiala Forest Data/SMEAR II Hyytiala Forest.xlsx",
        "Jasmine OseiEnin Lab Notebook/Response factors.xlsx",
        "Jasmine OseiEnin Lab Notebook/2023-2024/Summer 24/8mix_calc.xlsx",
        "Jasmine OseiEnin Lab Notebook/2023-2024/Summer 24/std_tidy.xlsx",
        "Lucia Labnotebooks/lab nootbook2022_monoterpene.pdf",
        "Lucia Labnotebooks/Labnotebook 2023_sesquiterpene.pdf",
        "Jasmine OseiEnin Lab Notebook/Coding commands.xlsx"
    ]

    docx_reader = DOCXReader()
    excel_reader = ExcelReader()
    pdf_reader = PDFReader() # Instantiate PDFReader

    for rel_path_str in files_to_inspect_rel_paths:
        file_path = base_notebooks_path / rel_path_str
        
        console.print(f"\n[bold yellow]--- Inspecting: {file_path.name} ---[/bold yellow]")
        
        if not file_path.exists():
            console.print(f"[red]File not found at {file_path}[/red]")
            continue

        try:
            if file_path.suffix.lower() == '.docx':
                # Summarize DOCX
                paragraphs = docx_reader.read_paragraphs(file_path)
                headings = docx_reader.read_headings(file_path)
                
                console.print("\n[bold blue]  First few paragraphs:[/bold blue]")
                if paragraphs:
                    for i, para_text in enumerate(paragraphs[:5]):
                        text = para_text.strip()
                        if text: console.print(f"  - {text}")
                else: console.print("  (No paragraphs found)")
                
                console.print("\n[bold blue]  First few headings:[/bold blue]")
                if headings:
                    for i, (level, heading_text) in enumerate(headings[:3]):
                        text = heading_text.strip()
                        if text: console.print(f"  - {text} (Level: {level})")
                else: console.print("  (No explicit headings found)")
                
                # Determine Relevance
                if "metabolomics sop" in file_path.name.lower() or "protocol" in file_path.name.lower():
                    console.print(f"[green]  Relevance: Highly relevant SOP/protocol for sample prep or collection, crucial for understanding data generation for your tree data and instrument readers.[/green]")
                else:
                     console.print(f"[green]  Relevance: Content suggests general lab protocol, good for reference.[/green]")


            elif file_path.suffix.lower() in ('.xlsx', '.xls'):
                # Summarize Excel
                import openpyxl # For listing sheets
                import polars as pl # For reading dataframes
                wb = openpyxl.load_workbook(file_path, read_only=True)
                sheets = wb.sheetnames
                wb.close()
                
                console.print(f"  [bold blue]Sheets found:[/bold blue] {sheets}")

                for sheet_name in sheets[:2]: # Inspect first 2 sheets
                    console.print(f"\n  [bold blue]--- Sheet: {sheet_name} ---[/bold blue]")
                    try:
                        df = pl.read_excel(source=file_path, sheet_name=sheet_name, engine="openpyxl", has_header=True)
                        if df.height > 0:
                            console.print(f"    [bold cyan]Columns:[/bold cyan] {df.columns}")
                            console.print("    [bold cyan]First 2 rows:[/bold cyan]")
                            for row in df.head(2).iter_rows(named=True):
                                console.print(f"      {row}")
                        else: console.print("    (Empty sheet or no data rows)")
                    except Exception as e:
                        console.print(f"    [red]Error reading sheet '{sheet_name}' with headers: {e}[/red]")
                        # Try reading as raw if header reading fails
                        try:
                            df = pl.read_excel(source=file_path, sheet_name=sheet_name, engine="openpyxl", has_header=False)
                            if df.height > 0:
                                console.print(f"    [bold cyan]Raw data (first row):[/bold cyan] {df.row(0)}")
                            else: console.print("    (Empty sheet or no data rows)")
                        except Exception as e_raw:
                            console.print(f"    [red]Error reading sheet '{sheet_name}' as raw: {e_raw}[/red]")
                
                # Determine Relevance
                if "response factors" in file_path.name.lower() or "calc" in file_path.name.lower() or "std" in file_path.name.lower():
                    console.print(f"[green]  Relevance: Contains calibration data, standard calculations, or response factors; highly relevant for your instrument data processing.[/green]")
                elif "gcms" in file_path.name.lower() or "smea" in file_path.name.lower() or "forest" in file_path.name.lower():
                     console.print(f"[green]  Relevance: Likely raw or processed GC-MS/environmental data; excellent for testing your `instrument_io` readers.[/green]")
                elif "coding" in file_path.name.lower():
                    console.print(f"[green]  Relevance: Contains coding commands/snippets; useful reference for your development work.[/green]")
                else:
                    console.print(f"[green]  Relevance: General Excel data, potentially useful for context.[/green]")


            elif file_path.suffix.lower() == '.pdf':
                # Summarize PDF
                text_content = pdf_reader.read_text(file_path) # Correct method
                
                console.print("\n[bold blue]  First few lines of text:[/bold blue]")
                if text_content:
                    lines = text_content.split('\n')
                    for line in lines[:7]: # Print first 7 lines
                        text = line.strip()
                        if text: console.print(f"  - {text}")
                else: console.print("  (No text content found or PDF is image-based)")

                # Determine Relevance
                if "monoterpene" in file_path.name.lower() or "sesquiterpene" in file_path.name.lower():
                    console.print(f"[green]  Relevance: Contains specific chemical class information (monoterpenes/sesquiterpenes); highly relevant if your `tree data` involves these compounds.[/green]")
                else:
                    console.print(f"[green]  Relevance: General PDF document, potentially useful for context.[/green]")

            else:
                console.print("[red]  File type not supported for detailed summary.[/red]")
                
        except Exception as e:
            console.print(f"[red]Error inspecting {file_path.name}: {e}[/red]")

if __name__ == "__main__":
    summarize_documents_at_once()
