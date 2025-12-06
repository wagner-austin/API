
import sys
from pathlib import Path
from openpyxl import load_workbook

def list_sheets():
    file_path = Path('C:/Users/austi/PROJECTS/UC Irvine/Celia Louise Braun Faiola - FaiolaLab/Important Docs/Chemical Inventory/02252021-Chemical Inventory.xlsx')
    
    try:
        wb = load_workbook(file_path, read_only=True)
        print(f"File: {file_path.name}")
        print("Sheets found:")
        for sheet in wb.sheetnames:
            print(f"- {sheet}")
    except Exception as e:
        print(f"Error reading file: {e}")

if __name__ == '__main__':
    list_sheets()
